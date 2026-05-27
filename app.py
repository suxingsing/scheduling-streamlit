import streamlit as st
from datetime import datetime, timedelta
import pandas as pd
import io
import math
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ------------------------------
# 一、全局核心配置
# ------------------------------
# 制程配置
PROCESS_CONFIG = {
    "SMT": {"default_uph": 1200, "default_lead_time": 3, "default_work_hours": 12, "desc": "贴片制程"},
    "预组": {"default_uph": 600, "default_lead_time": 2, "default_work_hours": 10, "desc": "预组装制程"},
    "组装": {"default_uph": 400, "default_lead_time": 2, "default_work_hours": 10, "desc": "整机组装制程"},
    "二测": {"default_uph": 800, "default_lead_time": 1, "default_work_hours": 10, "desc": "二次测试制程"},
    "包装": {"default_uph": 1500, "default_lead_time": 1, "default_work_hours": 10, "desc": "成品包装制程"}
}

# 爬坡斜率数据（DAY1到DAYn，产能从低到高递增，倒排时严格从DAY1开始推算）
RAMP_DATA = {
    "高端机": {
        "100%新人": [10, 20, 30, 40, 50, 60, 65, 70, 75, 80, 85, 90, 95, 95, 100],
        "50%新人": [15, 25, 35, 45, 55, 65, 75, 80, 85, 90, 95, 100],
        "0%新人": [20, 30, 40, 50, 60, 70, 80, 90, 95, 100]
    },
    "中端机": {
        "100%新人": [10, 20, 30, 40, 50, 60, 70, 80, 85, 90, 95, 100],
        "50%新人": [15, 30, 40, 50, 60, 70, 80, 90, 100],
        "0%新人": [20, 40, 60, 70, 80, 90, 100]
    },
    "低端机": {
        "100%新人": [20, 30, 40, 50, 60, 70, 80, 90, 95, 100],
        "50%新人": [30, 50, 70, 80, 85, 90, 95, 100],
        "0%新人": [40, 60, 80, 90, 95, 100]
    }
}

# ------------------------------
# 二、核心工具函数
# ------------------------------
def generate_full_date_list(start_date, end_date):
    """生成完整的自然日列表"""
    date_list = []
    current_date = start_date
    while current_date <= end_date:
        date_list.append(current_date)
        current_date += timedelta(days=1)
    return date_list

def get_previous_workday(base_date, days_to_go_back, rest_dates_set):
    """从基准日期往前数N个工作日，返回目标日期"""
    current_date = base_date
    count = 0
    while count < days_to_go_back:
        current_date -= timedelta(days=1)
        if current_date not in rest_dates_set:
            count += 1
    return current_date

def short_date_label(d):
    return f"{d.month}月{d.day}日"

def parse_template_date(value, fallback_year):
    """兼容模板里的 4月1日、4/1、2026-04-01 等日期写法。"""
    if pd.isna(value):
        return None
    if hasattr(value, "date"):
        return value.date()
    if isinstance(value, (int, float)):
        try:
            return pd.to_datetime(value, unit="D", origin="1899-12-30").date()
        except Exception:
            return None

    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d", "%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            year = parsed.year if "%Y" in fmt else fallback_year
            return parsed.replace(year=year).date()
        except ValueError:
            pass
    if "月" in text and "日" in text:
        try:
            month = int(text.split("月", 1)[0])
            day = int(text.split("月", 1)[1].split("日", 1)[0])
            return datetime(fallback_year, month, day).date()
        except Exception:
            return None
    return None

def build_material_template(start_date, end_date, default_initial_stock):
    """生成横排日期物料交期模板，日期范围跟随页面输入。"""
    date_list = generate_full_date_list(start_date, end_date)
    wb = Workbook()
    ws = wb.active
    ws.title = "物料交期输入"

    ws["A1"] = "物料交期"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, len(date_list) + 1))
    ws["A3"] = "物料期初库存"
    ws["A4"] = int(default_initial_stock)

    header_fill = PatternFill("solid", fgColor="C9B8AA")
    input_fill = PatternFill("solid", fgColor="F2DED2")
    value_fill = PatternFill("solid", fgColor="FFFFFF")
    border = Border(
        left=Side(style="thin", color="B7B7B7"),
        right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7"),
    )

    for col_idx, d in enumerate(date_list, start=2):
        ws.cell(row=3, column=col_idx, value=short_date_label(d))
        ws.cell(row=4, column=col_idx, value=0)

    for row_idx in range(3, 5):
        for col_idx in range(1, len(date_list) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if row_idx == 3:
                cell.fill = header_fill
                cell.font = Font(bold=True)
            elif col_idx == 1:
                cell.fill = input_fill
                cell.font = Font(bold=True)
            else:
                cell.fill = value_fill

    ws.freeze_panes = "B4"
    ws.column_dimensions["A"].width = 18
    for col_idx in range(2, len(date_list) + 2):
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = 13

    help_ws = wb.create_sheet("填写说明")
    help_ws.append(["字段", "填写要求"])
    help_ws.append(["物料期初库存", "填写排程开始前可用物料数量，位于 A4。"])
    help_ws.append(["日期", "第 3 行由系统按页面排产开始日期和需求最终截止日期横向生成。"])
    help_ws.append(["预计到料数量", "第 4 行填写每天预计到料数量，空白按 0 处理。"])
    help_ws.append(["到料规则", "T 日到料最早 T+1 日投入排产。"])
    help_ws.column_dimensions["A"].width = 18
    help_ws.column_dimensions["B"].width = 58

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def parse_material_upload(uploaded_file, start_date, end_date, default_initial_stock):
    """读取竖排模板；同时兼容旧横排模板和普通两列表。"""
    fallback_year = start_date.year
    uploaded_file.seek(0)
    raw_df = pd.read_excel(uploaded_file, sheet_name=0, header=None)

    header_row_idx = None
    for idx, row in raw_df.iterrows():
        first_cell = row.iloc[0] if len(row) else None
        if isinstance(first_cell, str) and "物料期初库存" in first_cell:
            header_row_idx = idx
            break

    if header_row_idx is not None and header_row_idx + 1 < len(raw_df):
        first_value = raw_df.iloc[header_row_idx, 1] if raw_df.shape[1] > 1 else None
        next_first = raw_df.iloc[header_row_idx + 1, 0]
        date_header_idx = header_row_idx + 1
        if isinstance(next_first, str) and "半成品" in next_first:
            date_header_idx = header_row_idx + 2
            next_first = raw_df.iloc[date_header_idx, 0] if date_header_idx < len(raw_df) else None

        if isinstance(next_first, str) and "日期" in next_first:
            material_initial_stock = default_initial_stock if pd.isna(first_value) else int(first_value)
            records = []
            invalid_cells = 0
            for row_idx in range(date_header_idx + 1, len(raw_df)):
                d = parse_template_date(raw_df.iloc[row_idx, 0], fallback_year)
                if d is None:
                    continue
                qty_value = raw_df.iloc[row_idx, 1] if raw_df.shape[1] > 1 else 0
                if pd.isna(qty_value) or qty_value == "":
                    qty = 0
                else:
                    try:
                        qty = int(qty_value)
                        if qty < 0:
                            raise ValueError
                    except Exception:
                        invalid_cells += 1
                        qty = 0
                records.append({"日期": d, "预计到料数量": qty})
            messages = []
            if invalid_cells:
                messages.append(f"有 {invalid_cells} 个到料数量无法识别，已按 0 处理。")
            return material_initial_stock, pd.DataFrame(records), messages

        date_row = raw_df.iloc[header_row_idx]
        value_row = raw_df.iloc[header_row_idx + 1]
        raw_initial_stock = value_row.iloc[0]
        material_initial_stock = default_initial_stock if pd.isna(raw_initial_stock) else int(raw_initial_stock)
        records = []
        invalid_cells = 0
        for col_idx in range(1, len(date_row)):
            d = parse_template_date(date_row.iloc[col_idx], fallback_year)
            if d is None:
                continue
            qty_value = value_row.iloc[col_idx] if col_idx < len(value_row) else 0
            if pd.isna(qty_value) or qty_value == "":
                qty = 0
            else:
                try:
                    qty = int(qty_value)
                    if qty < 0:
                        raise ValueError
                except Exception:
                    invalid_cells += 1
                    qty = 0
            records.append({"日期": d, "预计到料数量": qty})
        messages = []
        if invalid_cells:
            messages.append(f"有 {invalid_cells} 个到料数量无法识别，已按 0 处理。")
        return material_initial_stock, pd.DataFrame(records), messages

    uploaded_file.seek(0)
    table_df = pd.read_excel(uploaded_file)
    if {"日期", "预计到料数量"}.issubset(table_df.columns):
        material_initial_stock = default_initial_stock
        if "物料期初库存" in table_df.columns and not table_df["物料期初库存"].dropna().empty:
            material_initial_stock = int(table_df["物料期初库存"].dropna().iloc[0])
        material_df = table_df[["日期", "预计到料数量"]].copy()
        material_df["日期"] = pd.to_datetime(material_df["日期"]).dt.date
        material_df["预计到料数量"] = material_df["预计到料数量"].fillna(0).astype(int)
        return material_initial_stock, material_df, []

    raise ValueError("未找到“物料期初库存 + 日期列”的模板结构，也未找到“日期/预计到料数量”两列。")

def parse_material_plan(material_plan_df):
    plan = {}
    invalid_rows = 0
    if material_plan_df is None:
        return plan, invalid_rows
    for _, row in material_plan_df.iterrows():
        raw_date = row.get("日期")
        raw_qty = row.get("预计到料数量", row.get("物料交期"))
        if pd.isna(raw_date) and pd.isna(raw_qty):
            continue
        try:
            d = raw_date.date() if hasattr(raw_date, "date") else pd.to_datetime(raw_date).date()
            qty = 0 if pd.isna(raw_qty) else int(raw_qty)
        except Exception:
            invalid_rows += 1
            continue
        if qty < 0:
            invalid_rows += 1
            continue
        plan[d] = plan.get(d, 0) + qty
    return plan, invalid_rows

def build_work_hours_template(start_date, end_date, default_work_hours):
    """生成竖排日期单日工时模板，日期范围跟随页面输入。"""
    date_list = generate_full_date_list(start_date, end_date)
    wb = Workbook()
    ws = wb.active
    ws.title = "单日工时输入"

    ws["A1"] = "单日工时"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:B1")
    ws["A3"] = "日期"
    ws["B3"] = "单班组单日工时"

    header_fill = PatternFill("solid", fgColor="C9B8AA")
    input_fill = PatternFill("solid", fgColor="F2DED2")
    value_fill = PatternFill("solid", fgColor="FFFFFF")
    border = Border(
        left=Side(style="thin", color="B7B7B7"),
        right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7"),
    )

    for row_idx, d in enumerate(date_list, start=4):
        ws.cell(row=row_idx, column=1, value=short_date_label(d))
        ws.cell(row=row_idx, column=2, value=int(default_work_hours))

    for row_idx in range(3, len(date_list) + 4):
        for col_idx in range(1, 3):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if row_idx == 3:
                cell.fill = header_fill
                cell.font = Font(bold=True)
            elif col_idx == 1:
                cell.fill = input_fill
                cell.font = Font(bold=True)
            else:
                cell.fill = value_fill

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18

    help_ws = wb.create_sheet("填写说明")
    help_ws.append(["字段", "填写要求"])
    help_ws.append(["日期", "由系统按页面排产开始日期和需求最终截止日期自动生成。"])
    help_ws.append(["单班组单日工时", "填写每天单班组工作小时，UPH保持不变，空白沿用默认工时。"])
    help_ws.column_dimensions["A"].width = 18
    help_ws.column_dimensions["B"].width = 58

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def parse_work_hours_upload(uploaded_file, start_date, end_date, default_work_hours):
    fallback_year = start_date.year
    uploaded_file.seek(0)
    raw_df = pd.read_excel(uploaded_file, sheet_name=0, header=None)

    header_row_idx = None
    for idx, row in raw_df.iterrows():
        first_cell = row.iloc[0] if len(row) else None
        second_cell = row.iloc[1] if len(row) > 1 else None
        if isinstance(first_cell, str) and "日期" in first_cell and isinstance(second_cell, str) and "工时" in second_cell:
            header_row_idx = idx
            break

    if header_row_idx is not None:
        records = []
        invalid_cells = 0
        for row_idx in range(header_row_idx + 1, len(raw_df)):
            d = parse_template_date(raw_df.iloc[row_idx, 0], fallback_year)
            if d is None:
                continue
            hour_value = raw_df.iloc[row_idx, 1] if raw_df.shape[1] > 1 else default_work_hours
            if pd.isna(hour_value) or hour_value == "":
                hours = int(default_work_hours)
            else:
                try:
                    hours = int(hour_value)
                    if hours < 0 or hours > 24:
                        raise ValueError
                except Exception:
                    invalid_cells += 1
                    hours = int(default_work_hours)
            records.append({"日期": d, "单班组单日工时": hours})
        messages = []
        if invalid_cells:
            messages.append(f"有 {invalid_cells} 个工时无法识别，已沿用默认工时。")
        return pd.DataFrame(records), messages

    uploaded_file.seek(0)
    table_df = pd.read_excel(uploaded_file)
    if {"日期", "单班组单日工时"}.issubset(table_df.columns):
        hours_df = table_df[["日期", "单班组单日工时"]].copy()
        hours_df["日期"] = pd.to_datetime(hours_df["日期"]).dt.date
        hours_df["单班组单日工时"] = hours_df["单班组单日工时"].fillna(default_work_hours).astype(int).clip(0, 24)
        return hours_df, []

    raise ValueError("未找到“日期 / 单班组单日工时”两列。")

def parse_work_hours_plan(work_hours_plan_df, default_work_hours):
    plan = {}
    invalid_rows = 0
    if work_hours_plan_df is None:
        return plan, invalid_rows
    for _, row in work_hours_plan_df.iterrows():
        raw_date = row.get("日期")
        raw_hours = row.get("单班组单日工时", row.get("工时"))
        if pd.isna(raw_date) and pd.isna(raw_hours):
            continue
        try:
            d = raw_date.date() if hasattr(raw_date, "date") else pd.to_datetime(raw_date).date()
            hours = default_work_hours if pd.isna(raw_hours) else int(raw_hours)
            if hours < 0 or hours > 24:
                raise ValueError
        except Exception:
            invalid_rows += 1
            continue
        plan[d] = hours
    return plan, invalid_rows

def is_shift_empty(shift_daily_prod):
    """判断一个班组是否全月为空，用于隐藏空班组"""
    for val in shift_daily_prod:
        if isinstance(val, (int, float)) and val > 0:
            return False
    return True

def calculate_shift_total_production(shift_daily_prod):
    """计算一个班组的总产量，用于优化分析"""
    total = 0
    for val in shift_daily_prod:
        if isinstance(val, (int, float)):
            total += val
    return total

def to_int_or_zero(value):
    try:
        if pd.isna(value):
            return 0
        numeric_value = pd.to_numeric(value, errors="coerce")
        if pd.isna(numeric_value):
            return 0
        return int(numeric_value)
    except Exception:
        return 0

def build_schedule_insights(
    schedule_df,
    production_target,
    total_demand,
    initial_stock,
    special_occupy,
    total_exist_capacity,
    existing_shift_count,
    uph_base,
    material_schedule_enabled,
    material_warnings,
    mode,
):
    """基于排产结果表生成结论和建议，避免只展示明细表。"""
    if schedule_df.empty:
        return {}, [], []

    date_cols = list(schedule_df.columns[2:])
    shift_rows = schedule_df[schedule_df["班组/指标"].astype(str).str.startswith("班组", na=False)]
    old_shift_rows = shift_rows[shift_rows["班组/指标"].astype(str).str.contains("老班组", na=False)]
    new_shift_rows = shift_rows[shift_rows["班组/指标"].astype(str).str.contains("新班组", na=False)]

    daily_totals = []
    for col in date_cols:
        daily_totals.append(sum(to_int_or_zero(row[col]) for _, row in shift_rows.iterrows()))
    planned_output = sum(daily_totals)
    shortage = max(int(production_target) - planned_output, 0)
    surplus = max(planned_output - int(production_target), 0)

    last_prod_day = "无"
    for col, qty in zip(date_cols, daily_totals):
        if qty > 0:
            last_prod_day = col.replace("\n", " ")

    hours_rows = schedule_df[schedule_df["班组/指标"].astype(str) == "单班组单日工时"]
    if not hours_rows.empty:
        hours_row = hours_rows.iloc[0]
        daily_capacity_map = {col: to_int_or_zero(hours_row[col]) * int(uph_base) for col in date_cols}
    else:
        daily_capacity_map = {col: 0 for col in date_cols}

    old_idle_days = 0
    for col in date_cols:
        day_capacity = daily_capacity_map.get(col, 0)
        # 放空只统计非休息日且有排产工时的日期；休息日或 0 工时日期不计入放空。
        if day_capacity <= 0:
            continue
        day_is_idle = False
        for _, row in old_shift_rows.iterrows():
            val = row[col]
            produced_qty = to_int_or_zero(val)
            if str(val) == "当日放空" or produced_qty < day_capacity:
                day_is_idle = True
                break
        if day_is_idle:
            old_idle_days += 1

    material_gap_min = None
    if material_schedule_enabled and "累计物料gap" in schedule_df["班组/指标"].astype(str).values:
        gap_row = schedule_df[schedule_df["班组/指标"].astype(str) == "累计物料gap"].iloc[0]
        material_gap_values = [to_int_or_zero(gap_row[col]) for col in date_cols]
        material_gap_min = min(material_gap_values) if material_gap_values else None

    summary = {
        "planned_output": planned_output,
        "shortage": shortage,
        "surplus": surplus,
        "old_shift_count": len(old_shift_rows),
        "new_shift_count": len(new_shift_rows),
        "last_prod_day": last_prod_day,
        "old_idle_days": old_idle_days,
        "material_gap_min": material_gap_min,
    }

    conclusions = []
    if shortage > 0:
        conclusions.append(f"本次排产未完全覆盖目标，还差 {shortage:,} 件。")
    else:
        conclusions.append(f"本次排产可覆盖目标，计划产出 {planned_output:,} 件，目标 {int(production_target):,} 件。")
        if surplus > 0:
            conclusions.append(f"排产结果存在 {surplus:,} 件余量，主要来自班组产能粒度或爬坡排产边界。")

    if len(new_shift_rows) > 0:
        conclusions.append(f"系统启用了 {len(new_shift_rows)} 个新班组，新班组按爬坡曲线进行连续倒排。")
    else:
        conclusions.append(f"当前方案仅使用老班组，启用老班组 {len(old_shift_rows)} 个。")

    if old_idle_days > 0:
        conclusions.append(f"老班组存在 {old_idle_days} 个放空/未排满日期；统计时已排除休息日及 0 工时日期。")

    if material_schedule_enabled:
        conclusions.append("本次已按上传的物料交期约束排产，物料到料按 T+1 可用处理。")
    else:
        conclusions.append("本次未启用物料交期约束，排产结果仅按产能、工时、日历与爬坡规则计算。")

    suggestions = []
    if shortage > 0:
        suggestions.append("优先补充物料到料、增加可用工时或延长排产周期，然后重新排产。")
    elif "新班组" in mode:
        suggestions.append("建议复核新增班组的人员到位时间和爬坡比例，确认倒排启动日期可执行。")
    elif total_exist_capacity > production_target:
        suggestions.append("现有产能覆盖需求，可关注是否存在可减少班组或降低加班的空间。")

    if old_idle_days > 7:
        suggestions.append("老班组放空天数超过 7 天，建议优先采用新班组连续倒排方案，减少非连续生产带来的组织成本。")
    elif old_idle_days > 0:
        suggestions.append("老班组存在少量放空，可结合现场换线、人员调配或日工时调整进一步平滑产能。")

    if material_schedule_enabled and material_gap_min is not None and material_gap_min <= 0:
        suggestions.append("物料 gap 已接近 0，建议复核关键日期前一日到料，避免实际到料延迟导致排产中断。")
    if not material_schedule_enabled:
        suggestions.append("正式使用前建议上传物料交期 Excel，避免产能计划与真实到料节奏脱节。")

    for warning in material_warnings:
        if warning not in suggestions:
            suggestions.append(warning)

    return summary, conclusions, suggestions

def calculate_ramp_need_days(final_gap, ramp_curve, single_shift_daily):
    """
    核心倒排函数：根据缺口计算需要的爬坡天数和产能序列
    规则：从DAY1开始累加爬坡产能，直到覆盖缺口，返回需要的天数和产能列表
    """
    if final_gap <= 0:
        return 0, []

    cumulative_prod = 0
    need_days = 0
    prod_list = []

    # 从DAY1开始累加，直到覆盖缺口
    for rate in ramp_curve:
        if cumulative_prod >= final_gap:
            break
        daily_prod = int(single_shift_daily * (rate / 100.0))
        if cumulative_prod + daily_prod > final_gap:
            daily_prod = final_gap - cumulative_prod
        prod_list.append(daily_prod)
        cumulative_prod += daily_prod
        need_days += 1

    # 如果整个爬坡周期都不够，补满产天数
    if cumulative_prod < final_gap:
        remaining_gap = final_gap - cumulative_prod
        full_prod_days = math.ceil(remaining_gap / single_shift_daily)
        for i in range(full_prod_days):
            daily_prod = single_shift_daily
            if cumulative_prod + daily_prod > final_gap:
                daily_prod = final_gap - cumulative_prod
            prod_list.append(daily_prod)
            cumulative_prod += daily_prod
            need_days += 1

    return need_days, prod_list

def analyze_overtime_optimization(
    final_shift_total, existing_shift_count, single_shift_daily,
    production_end_date, rest_dates_set, full_date_list
):
    """
    加班优化分析：针对最后一个新增班组的产量，判断是否可通过减少休息日完成
    """
    if final_shift_total <= 0:
        return False, "", []

    # 1. 筛选生产窗口期内的休息日（<=生产截止日）
    production_rest_dates = [d for d in rest_dates_set if d <= production_end_date and d in full_date_list]
    if len(production_rest_dates) == 0:
        return False, "", []

    # 2. 计算单个休息日改成工作日，现有班组能增加的总产能
    single_rest_day_capacity = existing_shift_count * single_shift_daily
    # 3. 计算需要最少多少个休息日
    need_days = math.ceil(final_shift_total / single_rest_day_capacity)

    if need_days > len(production_rest_dates):
        return False, "", []

    # 4. 优先推荐周六（先改周六，保留周日休息），按日期从晚到早排序
    saturday_rest_dates = [d for d in production_rest_dates if d.weekday() == 5]
    other_rest_dates = [d for d in production_rest_dates if d.weekday() != 5]
    sorted_rest_dates = sorted(saturday_rest_dates, reverse=True) + sorted(other_rest_dates, reverse=True)

    # 5. 取需要的日期
    selected_dates = sorted_rest_dates[:need_days]
    total_add_capacity = need_days * single_rest_day_capacity

    # 6. 生成建议文本（已删除：推荐修改的休息日）
    suggest_text = f"""
⚠️ 加班优化建议：当前最后一个新增班组的总产量，可通过现有班组减少休息日（加班）完成，无需新增该班组！
🔹 最后一个新增班组总产量：{final_shift_total:,}
🔹 单个休息日改为工作日，现有班组可增加产能：{single_rest_day_capacity:,}（现有{existing_shift_count}个班组满产）
🔹 最少需要减少休息日数量：{need_days}天
🔹 修改后可增加总产能：{total_add_capacity:,}，完全覆盖该班组产量
🔹 重要说明：减少休息日重新排产，依然完全遵从现有正排/倒排规则、爬坡规则，不会打乱原有逻辑
"""
    return True, suggest_text, selected_dates

# ------------------------------
# 三、核心排产引擎
# ------------------------------
def schedule_engine(
    process_name,
    total_demand, initial_stock, special_occupy,
    uph_base, work_hours, existing_shift_count,
    schedule_start_date, demand_end_date, lead_time_days,
    rest_dates_set,
    material_initial_stock, material_plan_df,
    selected_model, new_human_ratio,
    material_enabled=True,
    work_hours_plan_df=None
):
    # 生产总目标
    production_target = max(total_demand + special_occupy - initial_stock, 0)
    default_single_shift_daily = uph_base * work_hours

    if production_target == 0:
        return pd.DataFrame(), "生产目标为0，无需排产", "", production_target, 0, default_single_shift_daily, 0, demand_end_date, 0, []

    # 生产截止日 = 需求截止日 - Lead Time 工作日
    production_end_date = get_previous_workday(demand_end_date, lead_time_days, rest_dates_set)
    full_date_list = generate_full_date_list(schedule_start_date, demand_end_date)
    total_days = len(full_date_list)
    if total_days == 0:
        return pd.DataFrame(), "错误：目标日期范围内无有效日期", "", production_target, 0, default_single_shift_daily, 0, demand_end_date, 0, []

    date_workday_flag = [d not in rest_dates_set for d in full_date_list]
    date_in_production_flag = [d <= production_end_date for d in full_date_list]
    date_to_idx = {d: i for i, d in enumerate(full_date_list)}
    total_workdays = sum(date_workday_flag)

    production_workday_indices = [i for i in range(total_days) if date_workday_flag[i] and date_in_production_flag[i]]
    reverse_production_workday_indices = production_workday_indices[::-1]
    production_workdays = len(production_workday_indices)

    if production_workdays == 0:
        return pd.DataFrame(), "错误：生产周期内无可用工作日", "", production_target, 0, default_single_shift_daily, 0, production_end_date, 0, []

    work_hours_plan, invalid_work_hours_rows = parse_work_hours_plan(work_hours_plan_df, int(work_hours))
    ignored_work_hours_dates = sorted(d for d in work_hours_plan if d not in set(full_date_list))
    daily_work_hours = [
        int(work_hours_plan.get(d, work_hours)) if date_workday_flag[idx] and date_in_production_flag[idx] else 0
        for idx, d in enumerate(full_date_list)
    ]
    daily_shift_capacity = [int(uph_base * hours) for hours in daily_work_hours]

    if material_enabled:
        material_plan, invalid_material_rows = parse_material_plan(material_plan_df)
        material_arrivals = [int(material_plan.get(d, 0)) for d in full_date_list]
        ignored_material_dates = sorted(d for d in material_plan if d not in set(full_date_list))
    else:
        invalid_material_rows = 0
        material_arrivals = [0] * total_days
        ignored_material_dates = []
    daily_scheduled = [0] * total_days

    def material_available_at_day_start(day_idx, scheduled_snapshot):
        if not material_enabled:
            return 10**18
        # 当日可用物料 = 前一天累计物料gap + 前一天预计到料数量。
        available_qty = int(material_initial_stock)
        for idx in range(day_idx):
            available_qty = available_qty - int(scheduled_snapshot[idx]) + int(material_arrivals[idx])
        return available_qty

    def available_material_for_day(day_idx):
        return max(0, material_available_at_day_start(day_idx, daily_scheduled) - int(daily_scheduled[day_idx]))

    def assign_production(shift, day_idx, desired_qty, remaining_qty, allow_partial=True):
        if allow_partial:
            actual_qty = min(int(desired_qty), int(remaining_qty), available_material_for_day(day_idx))
        elif int(remaining_qty) >= int(desired_qty) and available_material_for_day(day_idx) >= int(desired_qty):
            actual_qty = int(desired_qty)
        else:
            actual_qty = 0
        if actual_qty > 0:
            shift["daily_prod"][day_idx] = actual_qty
            daily_scheduled[day_idx] += actual_qty
        return actual_qty

    def init_shift(name, is_new, fill_idle=True):
        daily = []
        for idx in range(total_days):
            if fill_idle and date_workday_flag[idx] and date_in_production_flag[idx]:
                daily.append("当日放空")
            else:
                daily.append("")
        return {"name": name, "daily_prod": daily, "is_new": is_new}

    def count_old_shift_idle_days():
        idle_days = set()
        for shift in shifts_production:
            if shift["is_new"]:
                continue
            # 只遍历生产窗口内的工作日；休息日不会进入 production_workday_indices。
            for idx in production_workday_indices:
                day_capacity = int(daily_shift_capacity[idx])
                # 自定义为 0 工时的日期视同不可排产，不计入放空。
                if day_capacity <= 0:
                    continue
                val = shift["daily_prod"][idx]
                # 当日标记放空，或实际产出低于当日满工时产能，均计为该日期放空/未排满。
                if val == "当日放空" or (isinstance(val, (int, float)) and int(val) < day_capacity):
                    idle_days.add(idx)
        return len(idle_days)

    def can_place_sequence(start_pos, prod_sequence, scheduled_snapshot):
        simulated = scheduled_snapshot[:]
        for offset, prod in enumerate(prod_sequence):
            seq_pos = start_pos + offset
            if seq_pos >= len(production_workday_indices):
                return False
            day_idx = production_workday_indices[seq_pos]
            available_qty = material_available_at_day_start(day_idx, simulated) - int(simulated[day_idx])
            if available_qty < int(prod):
                return False
            simulated[day_idx] += int(prod)
        return True

    def place_sequence(shift, start_pos, prod_sequence):
        produced = 0
        for offset, prod in enumerate(prod_sequence):
            day_idx = production_workday_indices[start_pos + offset]
            shift["daily_prod"][day_idx] = int(prod)
            daily_scheduled[day_idx] += int(prod)
            produced += int(prod)
        return produced

    def build_one_shift_sequence(target_qty, start_pos=0, max_days=None):
        ramp_curve = RAMP_DATA[selected_model][new_human_ratio]
        seq = []
        produced = 0
        max_days = production_workdays if max_days is None else int(max_days)
        for workday_seq in range(max_days):
            seq_pos = start_pos + workday_seq
            if seq_pos >= len(production_workday_indices):
                break
            day_idx = production_workday_indices[seq_pos]
            rate = ramp_curve[workday_seq] if workday_seq < len(ramp_curve) else 100
            qty = int(daily_shift_capacity[day_idx] * (rate / 100.0))
            qty = min(qty, int(target_qty) - produced)
            if qty <= 0:
                break
            seq.append(qty)
            produced += qty
            if produced >= int(target_qty):
                break
        return seq

    def calc_material_gap_row(scheduled_snapshot):
        gap_row = []
        for idx, qty in enumerate(scheduled_snapshot):
            gap_row.append(material_available_at_day_start(idx, scheduled_snapshot) - int(qty))
        return gap_row

    def calc_material_available_row(scheduled_snapshot):
        return [material_available_at_day_start(idx, scheduled_snapshot) for idx in range(total_days)]

    def add_new_reverse_shifts(target_qty, reset_existing=True):
        nonlocal daily_scheduled, shifts_production
        if reset_existing:
            daily_scheduled = [0] * total_days
            shifts_production = []
        remaining = int(target_qty)
        new_shift_total = 0
        shift_no = len(shifts_production) + 1

        while remaining > 0 and shift_no <= 50:
            best_start = None
            best_sequence = []
            best_total = 0
            for start_pos in range(len(production_workday_indices)):
                max_days = len(production_workday_indices) - start_pos
                prod_sequence = build_one_shift_sequence(remaining, start_pos=start_pos, max_days=max_days)
                if not prod_sequence:
                    continue
                produced = sum(prod_sequence)
                if produced <= 0:
                    continue
                if not can_place_sequence(start_pos, prod_sequence, daily_scheduled):
                    continue
                if produced > best_total or (produced == best_total and (best_start is None or start_pos > best_start)):
                    best_start = start_pos
                    best_sequence = prod_sequence
                    best_total = produced

            if best_start is None:
                break

            shift = init_shift(f"班组{shift_no}(新班组-连续倒排)", True, fill_idle=False)
            produced = place_sequence(shift, best_start, best_sequence)
            shifts_production.append(shift)
            remaining -= produced
            new_shift_total += produced
            shift_no += 1

        return remaining, new_shift_total

    def run_old_forward_schedule(old_shift_count, target_qty):
        nonlocal daily_scheduled
        daily_scheduled = [0] * total_days
        candidate_shifts = [init_shift(f"班组{i+1}(老班组)", False) for i in range(old_shift_count)]
        remaining = int(target_qty)

        for day_idx in production_workday_indices:
            if remaining <= 0:
                break
            for shift_idx in range(old_shift_count):
                if remaining <= 0:
                    break
                day_capacity = int(daily_shift_capacity[day_idx])
                prod = min(day_capacity, remaining)
                actual = assign_production(
                    candidate_shifts[shift_idx],
                    day_idx,
                    prod,
                    remaining,
                    allow_partial=(remaining < day_capacity),
                )
                remaining -= actual
                if actual == 0:
                    break

        return candidate_shifts, daily_scheduled[:], remaining

    # 产能计算
    capacity_1_shift_full = sum(int(daily_shift_capacity[idx]) for idx in production_workday_indices)
    total_exist_capacity = capacity_1_shift_full * existing_shift_count
    demand_gap = production_target - total_exist_capacity
    small_gap_threshold = 2 * default_single_shift_daily

    # ============================
    # 初始化班组：后续按实际启用班组数生成，避免产能过剩时保留多余班组
    # ============================
    shifts_production = []
    run_mode = ""

    # ============================
    # 场景1：产能过剩
    # ============================
    final_shift_total = 0
    old_material_gap_row = None
    if demand_gap <= 0:
        run_mode = "减少班组模式（产能过剩，正排逻辑）"
        selected_old_count = existing_shift_count
        selected_remaining = production_target
        selected_shifts = []
        selected_daily = [0] * total_days

        for candidate_count in range(1, existing_shift_count + 1):
            if candidate_count * capacity_1_shift_full < production_target:
                continue
            candidate_shifts, candidate_daily, candidate_remaining = run_old_forward_schedule(candidate_count, production_target)
            if candidate_remaining <= 0:
                selected_old_count = candidate_count
                selected_remaining = 0
                selected_shifts = candidate_shifts
                selected_daily = candidate_daily
                break
            if not selected_shifts or candidate_remaining < selected_remaining:
                selected_old_count = candidate_count
                selected_remaining = candidate_remaining
                selected_shifts = candidate_shifts
                selected_daily = candidate_daily

        if not selected_shifts:
            selected_shifts, selected_daily, selected_remaining = run_old_forward_schedule(existing_shift_count, production_target)

        shifts_production = selected_shifts
        daily_scheduled = selected_daily
        remaining_demand = selected_remaining

        old_material_gap_row = calc_material_gap_row(daily_scheduled)
        idle_days = count_old_shift_idle_days()
        if idle_days > 7:
            remaining_demand, final_shift_total = add_new_reverse_shifts(production_target)
            old_material_gap_row = None
            run_mode = "新班组连续倒排模式（老班组月度放空超过7天，自动切换）"
            message = f"✅ 排产完成 | {run_mode} | 老班组放空{idle_days}天，已切换为新班组倒排"
        else:
            if remaining_demand > 0:
                remaining_demand, final_shift_total = add_new_reverse_shifts(remaining_demand, reset_existing=False)
                run_mode = "老班组正排 + 新班组连续倒排模式"
                message = f"✅ 排产完成 | {run_mode} | 老班组放空{idle_days}天，按实际剩余需求启用新班组"
            else:
                reduced_count = existing_shift_count - selected_old_count
                message = f"✅ 排产完成 | {run_mode} | 启用老班组{selected_old_count}个，减少{reduced_count}个 | 老班组放空{idle_days}天"

    # ============================
    # 场景2：产能不足
    # ============================
    else:
        run_mode = "新增班组模式（产能不足，收尾班组严格遵循爬坡规则倒排）"
        shifts_production, daily_scheduled, remaining_demand = run_old_forward_schedule(existing_shift_count, production_target)

        final_shift_total = 0
        old_material_gap_row = calc_material_gap_row(daily_scheduled)
        idle_days = count_old_shift_idle_days()
        if idle_days > 7:
            remaining_demand, final_shift_total = add_new_reverse_shifts(production_target)
            old_material_gap_row = None
            run_mode = "新班组连续倒排模式（老班组月度放空超过7天，自动切换）"
            message = f"✅ 排产完成 | {run_mode} | 老班组放空{idle_days}天，已切换为新班组倒排"
        else:
            old_shift_count_after = len(shifts_production)
            remaining_demand, final_shift_total = add_new_reverse_shifts(remaining_demand, reset_existing=False) if remaining_demand > 0 else (0, 0)
            if old_shift_count_after and final_shift_total > 0:
                run_mode = "老班组正排 + 新班组连续倒排模式"
            message = f"✅ 排产完成 | {run_mode} | 启用班组总数：{len(shifts_production)}个 | 生产必须完成截止日：{production_end_date.month}月{production_end_date.day}日"

    # ============================
    # 统计行
    # ============================
    daily_total = [0] * total_days
    for day_idx in range(total_days):
        day_sum = 0
        for shift in shifts_production:
            val = shift["daily_prod"][day_idx]
            if isinstance(val, (int, float)):
                day_sum += int(val)
        daily_total[day_idx] = day_sum

    material_gap_row = []
    material_usable_row = []
    if material_enabled:
        for day_idx in range(total_days):
            available_qty = material_available_at_day_start(day_idx, daily_total)
            material_usable_row.append(int(available_qty))
            material_gap_row.append(int(available_qty) - int(daily_total[day_idx]))

    # 累计产量
    cumulative_row = [int(initial_stock)]
    current_cum = int(initial_stock)
    for val in daily_total:
        current_cum += val
        cumulative_row.append(current_cum)

    # 成品转化累计：Lead Time 按自然日偏移，T 日生产在 T+Lead Time 日转化。
    convert_row = [int(initial_stock)]
    convert_additions = [0] * total_days
    for prod_idx, qty in enumerate(daily_total):
        complete_idx = prod_idx + int(lead_time_days)
        if complete_idx < total_days:
            convert_additions[complete_idx] += int(qty)

    convert_cum = int(initial_stock)
    for day_idx in range(total_days):
        convert_cum += int(convert_additions[day_idx])
        convert_row.append(int(convert_cum))

    # 隐藏空班组
    final_shifts = []
    for shift in shifts_production:
        if not is_shift_empty(shift["daily_prod"]):
            final_shifts.append(shift)

    # 表格
    columns = ["班组/指标", "期初数据"]
    for d in full_date_list:
        week_name = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][d.weekday()]
        columns.append(f"{d.month}月{d.day}日\n{week_name}")

    rows = []
    rows.append(["单班组单日工时", ""] + daily_work_hours)
    if material_enabled:
        rows.append(["预计到料数量", int(material_initial_stock)] + material_arrivals)
        rows.append(["当日可用物料", int(material_initial_stock)] + material_usable_row)
    rows.append(["特殊占用（需生产）", int(special_occupy)] + [""] * total_days)
    old_shifts = [shift for shift in final_shifts if not shift["is_new"]]
    new_shifts = [shift for shift in final_shifts if shift["is_new"]]
    for shift in old_shifts:
        rows.append([shift["name"], ""] + shift["daily_prod"])
    if material_enabled and new_shifts and old_material_gap_row is not None:
        rows.append(["老班组后累计物料gap", ""] + old_material_gap_row)
    for shift in new_shifts:
        rows.append([shift["name"], ""] + shift["daily_prod"])
    if material_enabled:
        rows.append(["累计物料gap", int(material_initial_stock)] + material_gap_row)
    rows.append(["累计产量", int(initial_stock)] + cumulative_row[1:])
    rows.append(["成品转化累计", ""] + convert_row[1:])

    material_warnings = []
    if invalid_work_hours_rows:
        material_warnings.append(f"单日工时表中有 {invalid_work_hours_rows} 行无法识别，已沿用默认工时。")
    if ignored_work_hours_dates:
        preview = "、".join(f"{d.month}/{d.day}" for d in ignored_work_hours_dates[:5])
        material_warnings.append(f"有 {len(ignored_work_hours_dates)} 个工时日期不在排程周期内，已忽略：{preview}")
    if material_enabled and invalid_material_rows:
        material_warnings.append(f"物料交期表中有 {invalid_material_rows} 行无法识别，已忽略。")
    if material_enabled and ignored_material_dates:
        preview = "、".join(f"{d.month}/{d.day}" for d in ignored_material_dates[:5])
        material_warnings.append(f"有 {len(ignored_material_dates)} 个到料日期不在排程周期内，已忽略：{preview}")
    produced_total = sum(daily_total)
    if produced_total < production_target:
        shortage = production_target - produced_total
        material_warnings.append(f"当前物料交期约束下仍有 {shortage:,} 件未排完，请补充到料、增加产能或延长周期。")

    schedule_df = pd.DataFrame(rows, columns=columns)
    return schedule_df, message, run_mode, production_target, total_workdays, default_single_shift_daily, total_exist_capacity, production_end_date, final_shift_total, material_warnings

# ------------------------------
# 四、页面配置
# ------------------------------
st.set_page_config(page_title="智能排产系统", page_icon="📊", layout="wide")
st.markdown(
    """
    <style>
    .material-rule-panel {
        display: inline-block;
        max-width: 980px;
        padding: 16px 20px;
        border: 1px solid #dbe3ef;
        border-radius: 8px;
        background: #f7fbff;
        color: #1f3b57;
        line-height: 1.65;
        font-size: 16px;
    }
    .material-rule-panel ol {
        margin: 0;
        padding-left: 20px;
    }
    .material-rule-panel li {
        margin-bottom: 8px;
    }
    .material-rule-panel li:last-child {
        margin-bottom: 0;
    }
    .material-section-title {
        margin: 4px 0 10px 0;
        font-weight: 700;
        color: #2f3340;
    }
    div[data-testid="stDownloadButton"] > button {
        width: 100%;
        min-height: 72px;
        height: 72px;
    }
    div[data-testid="stFileUploader"] {
        width: 100%;
    }
    div[data-testid="stFileUploader"] section {
        width: 100%;
        min-height: 72px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)
st.title("智能排产系统")
st.divider()

# 制程选择
st.subheader("选择生产制程")
selected_process = st.selectbox(
    "请选择要排产的制程",
    options=list(PROCESS_CONFIG.keys()),
    index=2,
    format_func=lambda x: f"{x} - {PROCESS_CONFIG[x]['desc']}"
)
current_process_config = PROCESS_CONFIG[selected_process]
st.divider()

# 核心参数
st.subheader(f"【{selected_process}】制程核心参数配置")
target_col, date_col, process_col = st.columns([1.15, 1, 1])
with target_col:
    st.markdown("**需求与库存**")
    total_demand = st.number_input("总需求量", min_value=0, value=460000, step=1000)
    initial_stock = st.number_input("半成品期初库存", min_value=0, value=6000, step=1000)
    special_occupy = st.number_input("特殊占用（需生产）", min_value=0, value=0, step=100)
with date_col:
    st.markdown("**日期与班组**")
    existing_shift_count = st.number_input("现有老班组数量", min_value=1, max_value=10, value=3)
    schedule_start_date = st.date_input("排产开始日期", value=datetime(2026, 4, 1).date())
    demand_end_date = st.date_input("需求最终截止日期", value=datetime(2026, 4, 30).date())
with process_col:
    st.markdown("**制程能力**")
    lead_time_days = st.number_input("成品转化Lead Time(工作日)", min_value=0, value=current_process_config['default_lead_time'])
    uph_base = st.number_input("单班组UPH", min_value=0, value=current_process_config['default_uph'])
st.divider()

# 工厂日历
st.subheader("工厂日历设置")
default_sunday_rest = st.checkbox("默认周日自动设为休息日", value=True)
full_calendar_dates = generate_full_date_list(schedule_start_date, demand_end_date)
if default_sunday_rest:
    default_rest_dates = [d for d in full_calendar_dates if d.weekday() == 6]
else:
    default_rest_dates = []

rest_dates = st.multiselect(
    "自定义休息日",
    options=full_calendar_dates,
    default=default_rest_dates,
    format_func=lambda x: f"{x.month}月{x.day}日 周{['一','二','三','四','五','六','日'][x.weekday()]}"
)
rest_dates_set = set(rest_dates)
workday_count = sum(1 for d in full_calendar_dates if d not in rest_dates_set)
st.info(f"📅 周期共 {len(full_calendar_dates)} 天 | 工作日 {workday_count} 天 | 休息日 {len(rest_dates_set)} 天")
st.divider()

# 单日工时
st.subheader("单日工时输入")
st.caption("UPH保持不变；每日产能 = 单班组UPH × 当日单班组工时。休息日会自动按 0 工时处理。")
hours_mode_col, hours_default_col, hours_capacity_col = st.columns([1.2, 1, 1])
with hours_mode_col:
    work_hours_mode = st.radio(
        "工时输入方式",
        options=["统一工时", "按日期编辑"],
        horizontal=True,
    )
with hours_default_col:
    work_hours = st.number_input(
        "默认单班组单日工时",
        min_value=1,
        max_value=24,
        value=current_process_config["default_work_hours"],
        step=1,
    )
with hours_capacity_col:
    st.metric("默认单班组单日满产", f"{uph_base * work_hours:,}")

work_hours_plan_df = pd.DataFrame({
    "日期": full_calendar_dates,
    "是否休息日": ["是" if d in rest_dates_set else "否" for d in full_calendar_dates],
    "单班组单日工时": [0 if d in rest_dates_set else int(work_hours) for d in full_calendar_dates],
})
if work_hours_mode == "按日期编辑":
    edited_work_hours_df = st.data_editor(
        work_hours_plan_df,
        use_container_width=True,
        hide_index=True,
        height=260,
        disabled=["日期", "是否休息日"],
        column_config={
            "日期": st.column_config.DateColumn("日期", format="YYYY-MM-DD"),
            "是否休息日": st.column_config.TextColumn("是否休息日"),
            "单班组单日工时": st.column_config.NumberColumn("单班组单日工时", min_value=0, max_value=24, step=1),
        },
        key="work_hours_editor",
    )
    work_hours_plan_df = edited_work_hours_df[["日期", "单班组单日工时"]].copy()
else:
    st.info("当前按统一工时计算；如前期 8 小时、后期满产，可切换为“按日期编辑”直接修改每天工时。")
    work_hours_plan_df = work_hours_plan_df[["日期", "单班组单日工时"]].copy()
st.divider()

# 物料交期
st.subheader("物料交期输入")
material_enabled = st.checkbox("启用物料交期约束", value=True)
default_material_initial_stock = 0
material_initial_stock = 0
material_plan_df = pd.DataFrame({
    "日期": full_calendar_dates,
    "预计到料数量": [0] * len(full_calendar_dates),
})
material_schedule_enabled = False
if material_enabled:
    default_material_initial_stock = 14000
    material_initial_stock = int(default_material_initial_stock)

    material_template = build_material_template(
        schedule_start_date,
        demand_end_date,
        default_material_initial_stock,
    )
    template_col, upload_col = st.columns(2)
    with template_col:
        st.markdown("<div class='material-section-title'>模板下载</div>", unsafe_allow_html=True)
        st.download_button(
            "下载Excel模板",
            data=material_template,
            file_name=f"物料交期输入模板_{schedule_start_date.strftime('%Y%m%d')}_{demand_end_date.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with upload_col:
        st.markdown("<div class='material-section-title'>Excel上传</div>", unsafe_allow_html=True)
        uploaded_material_file = st.file_uploader(
            "上传已填写的物料交期Excel",
            type=["xlsx", "xls"],
            key="material_upload",
            label_visibility="collapsed",
        )

    if uploaded_material_file is not None:
        try:
            material_initial_stock, material_plan_df, upload_messages = parse_material_upload(
                uploaded_material_file,
                schedule_start_date,
                demand_end_date,
                default_material_initial_stock,
            )
            material_schedule_enabled = True
            st.success(f"已读取物料期初库存：{material_initial_stock:,}，到料日期 {len(material_plan_df)} 条。")
            for msg in upload_messages:
                st.warning(msg)
        except Exception as exc:
            st.error(f"物料交期Excel解析失败：{exc}")
            st.stop()

    preview_col, note_col = st.columns(2)
    with preview_col:
        st.markdown("<div class='material-section-title'>物料交期预览</div>", unsafe_allow_html=True)
        st.dataframe(
            material_plan_df,
            use_container_width=True,
            hide_index=True,
            height=260,
            column_config={
                "日期": st.column_config.DateColumn("日期", format="YYYY-MM-DD"),
                "预计到料数量": st.column_config.NumberColumn("预计到料数量", min_value=0, step=1000),
            },
        )
    with note_col:
        st.markdown("<div class='material-section-title'>到料规则</div>", unsafe_allow_html=True)
        st.markdown(
            """
            <div class="material-rule-panel">
                <ol>
                    <li>物料到厂后按 <strong>T+1</strong> 可用：T 日到料，最早从 T+1 日开始投入排产。</li>
                    <li>每日排产前先核算可用物料：<strong>当日可用物料 = 前一天物料 gap + 前一天预计到料数量</strong>。</li>
                </ol>
            </div>
            """,
            unsafe_allow_html=True,
        )
else:
    st.info("已关闭物料交期约束，排产不受物料到料限制。")
st.divider()

# 爬坡规则
st.subheader("爬坡规则配置（仅新增班组适用）")
col1, col2 = st.columns(2)
with col1:
    selected_model = st.selectbox("选择机型", options=list(RAMP_DATA.keys()), index=0)
    new_human_ratio = st.selectbox("选择新人占比", options=list(RAMP_DATA[selected_model].keys()), index=0)
with col2:
    current_ramp_curve = RAMP_DATA[selected_model][new_human_ratio]
    ramp_df = pd.DataFrame({
        "爬坡天数": [f"DAY{i+1}" for i in range(len(current_ramp_curve))],
        "产能比例": [f"{x}%" for x in current_ramp_curve]
    }).T
    st.dataframe(ramp_df, use_container_width=True, hide_index=True)
    st.caption(f"爬坡总天数：{len(current_ramp_curve)}天")
st.divider()

# 执行排产
st.subheader("排产结果")
if st.button(f"开始【{selected_process}】制程排产", type="primary", use_container_width=True):
    schedule_df, message, mode, production_target, total_workdays, single_shift_daily, total_exist_capacity, production_end_date, final_shift_total, material_warnings = schedule_engine(
        process_name=selected_process,
        total_demand=total_demand,
        initial_stock=initial_stock,
        special_occupy=special_occupy,
        uph_base=uph_base,
        work_hours=work_hours,
        existing_shift_count=existing_shift_count,
        schedule_start_date=schedule_start_date,
        demand_end_date=demand_end_date,
        lead_time_days=lead_time_days,
        rest_dates_set=rest_dates_set,
        material_initial_stock=material_initial_stock,
        material_plan_df=material_plan_df,
        selected_model=selected_model,
        new_human_ratio=new_human_ratio,
        material_enabled=material_schedule_enabled,
        work_hours_plan_df=work_hours_plan_df,
    )

    # 加班提示
    can_optimize, optimize_suggest, selected_dates = analyze_overtime_optimization(
        final_shift_total, existing_shift_count, single_shift_daily,
        production_end_date, rest_dates_set, full_calendar_dates
    )
    if can_optimize:
        st.warning(optimize_suggest)
    for msg in material_warnings:
        st.warning(msg)

    # 核心计算结果
    st.markdown("### 📊 核心计算结果")
    st.markdown(f"- **总需求量**：{total_demand:,}")
    st.markdown(f"- **半成品期初库存**：{initial_stock:,}")
    if material_schedule_enabled:
        st.markdown(f"- **物料期初库存**：{material_initial_stock:,}")
    else:
        st.markdown("- **物料交期约束**：未启用")
    st.markdown(f"- **特殊占用（需生产）**：{special_occupy:,}")
    st.markdown(f"- **最终生产目标**：{production_target:,}")
    st.markdown(f"- **现有班组总产能**：{total_exist_capacity:,}")
    st.markdown(f"- **周期有效工作日**：{total_workdays}天")
    st.markdown(f"- **需求最终截止日**：{demand_end_date.month}月{demand_end_date.day}日")
    st.markdown(f"- **生产必须完成截止日**：{production_end_date.month}月{production_end_date.day}日（提前{lead_time_days}个工作日）")
    st.markdown(f"- **当前运行模式**：{mode}")
    st.divider()

    if mode == "small_gap":
        st.warning(message)
    elif "错误" in message:
        st.error(message)
    else:
        st.success(message)
        insight_summary, conclusions, suggestions = build_schedule_insights(
            schedule_df=schedule_df,
            production_target=production_target,
            total_demand=total_demand,
            initial_stock=initial_stock,
            special_occupy=special_occupy,
            total_exist_capacity=total_exist_capacity,
            existing_shift_count=existing_shift_count,
            uph_base=uph_base,
            material_schedule_enabled=material_schedule_enabled,
            material_warnings=material_warnings,
            mode=mode,
        )
        if insight_summary:
            st.markdown("### 排产结论与建议")
            metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
            metric_col1.metric("计划产出", f"{insight_summary['planned_output']:,}")
            if insight_summary["shortage"] > 0:
                metric_col2.metric("目标差异", f"缺口 {insight_summary['shortage']:,}")
            else:
                metric_col2.metric("目标差异", f"余量 {insight_summary['surplus']:,}")
            metric_col3.metric(
                "启用班组",
                f"老{insight_summary['old_shift_count']} / 新{insight_summary['new_shift_count']}",
            )
            metric_col4.metric("最后生产日", insight_summary["last_prod_day"])

            if material_schedule_enabled and insight_summary["material_gap_min"] is not None:
                st.caption(f"最低物料 gap：{insight_summary['material_gap_min']:,}")

            conclusion_col, suggestion_col = st.columns(2)
            with conclusion_col:
                st.markdown("**结论**")
                for item in conclusions:
                    st.markdown(f"- {item}")
            with suggestion_col:
                st.markdown("**建议**")
                for item in suggestions:
                    st.markdown(f"- {item}")
            st.divider()

        st.dataframe(schedule_df, use_container_width=True, height=500)

        # Excel导出
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            schedule_df.to_excel(writer, sheet_name=f'{selected_process}排产表', index=False)
            work_hours_plan_df.to_excel(writer, sheet_name='单日工时输入', index=False)
            if material_schedule_enabled:
                material_plan_df.to_excel(writer, sheet_name='物料交期输入', index=False)
        st.download_button(
            label=f"下载【{selected_process}】制程排产表Excel",
            data=buffer.getvalue(),
            file_name=f"{selected_process}制程_智能排产计划表_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
